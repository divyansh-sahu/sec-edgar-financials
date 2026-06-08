"""
Run once (locally) to extract minimal tag data from all taxonomy Excel files
and write taxonomy/data.json. Commit data.json; the xlsx files stay local.

Usage:
    python taxonomy/build_data.py
"""

import json
import re
from pathlib import Path
import openpyxl

TAXONOMY_DIR = Path(__file__).parent
OUT = TAXONOMY_DIR / "data.json"

_YEAR_RE   = re.compile(r"(?<!\d)(\d{4})(?!\d)")
_SHEET_NAMES = ("Concepts", "Elements")


def _year_from_stem(stem: str) -> int | None:
    tokens = re.split(r"[^a-zA-Z0-9]", stem)
    if any(t.upper() == "SRT" for t in tokens):
        return None
    m = _YEAR_RE.search(stem)
    return int(m.group(1)) if m else None


def extract(path: Path) -> dict[str, dict]:
    """Return {tag_name: {prefix, filer_count, deprecated}} for one xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = None
    for name in _SHEET_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return {}

    tags = {}
    col = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            col = {h: idx for idx, h in enumerate(row) if h is not None}
            has_fc = "Filer Usage Count" in col
            continue

        prefix = row[col.get("prefix", -1)] if col.get("prefix") is not None else None
        if prefix not in ("us-gaap", "srt"):
            continue
        if row[col["abstract"]]:
            continue
        name = row[col["name"]]
        if not name:
            continue
        tags[name] = {
            "p":  prefix,
            "fc": int(row[col["Filer Usage Count"]] or 0) if has_fc else 0,
            "d":  bool(row[col["deprecatedDate"]]) if col.get("deprecatedDate") is not None else False,
        }

    wb.close()
    return tags


def main():
    data = {}
    for path in sorted(TAXONOMY_DIR.glob("*.xlsx")):
        year = _year_from_stem(path.stem)
        if year is None:
            continue
        print(f"  Processing {path.name} → year {year} ...", end=" ", flush=True)
        tags = extract(path)
        data[str(year)] = tags
        print(f"{len(tags)} tags")

    OUT.write_text(json.dumps(data, separators=(",", ":")), encoding="utf-8")
    size_mb = OUT.stat().st_size / 1_000_000
    print(f"\nWrote {OUT}  ({size_mb:.1f} MB, {len(data)} years)")


if __name__ == "__main__":
    main()
